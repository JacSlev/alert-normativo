import pytest
import os
import shutil
import openpyxl
from datetime import date
from output.excel_logger import append_news
from output.pptx_generator import generate_pptx

TEMPLATE_XLSX = "assets/Template_settimanale.xlsx"
TEMPLATE_PPTX = "assets/_CLEAN.pptx"

# Anno "odierno" per i filtri di publish (col L è popolata da append_news con date.today())
TODAY_ANNO = date.today().strftime("%Y")

SAMPLE_APPROVED = [
    {"categoria": "BANKING", "fonte": "EBA",
     "titolo": "EBA pubblica linee guida FRTB", "descrizione": "L'EBA ha pubblicato linee guida.",
     "data_originale": "14/05/2026", "url": "https://eba.eu/1", "includi_in_pptx": "SI"},
]


@pytest.fixture
def excel_path(tmp_path):
    if not os.path.exists(TEMPLATE_XLSX) or not os.path.exists(TEMPLATE_PPTX):
        pytest.skip("Template XLSX o PPTX non trovati in assets/")
    dest = str(tmp_path / "db.xlsx")
    shutil.copy(TEMPLATE_XLSX, dest)
    return dest


def _set_edition_columns(path: str, edizione: str, anno: str):
    """Compila col J (edizione) e col L (anno) per tutte le righe dati, come fa il responsabile."""
    wb = openpyxl.load_workbook(path)
    ws = wb["Monitoraggio finance"]
    for r in range(3, ws.max_row + 1):
        if ws.cell(r, 1).value is not None:
            ws.cell(r, 10).value = edizione   # column J
            ws.cell(r, 12).value = anno       # column L
    wb.save(path)


def test_generate_pptx_exits_on_zero_news(excel_path, tmp_path):
    """DB senza righe corrispondenti → SystemExit e nessuna PPTX creata."""
    out = str(tmp_path / "out.pptx")
    with pytest.raises(SystemExit):
        generate_pptx(TEMPLATE_PPTX, excel_path, out,
                      numero="1", mese="06", anno=TODAY_ANNO,
                      edizione_numero="1", anno_filtro=TODAY_ANNO)
    assert not os.path.exists(out)


def test_generate_pptx_happy_path(excel_path, tmp_path):
    """1 notizia approvata con J/L corrette → la PPTX viene creata."""
    append_news(excel_path, SAMPLE_APPROVED)
    _set_edition_columns(excel_path, edizione="1", anno=TODAY_ANNO)
    out = str(tmp_path / "out.pptx")
    generate_pptx(TEMPLATE_PPTX, excel_path, out,
                  numero="1", mese="06", anno=TODAY_ANNO,
                  edizione_numero="1", anno_filtro=TODAY_ANNO)
    assert os.path.exists(out)
