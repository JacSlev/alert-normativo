import pytest
import os
import shutil
import openpyxl
from datetime import date
from openpyxl import load_workbook
from output.excel_logger import ensure_excel_exists, count_existing_rows, append_news, get_output_path, read_approved_news, count_filter_stages

SAMPLE_NEWS = [
    {"categoria": "BANKING", "fonte": "EBA",
     "titolo": "EBA pubblica linee guida FRTB", "descrizione": "L'EBA ha pubblicato linee guida.",
     "data_originale": "14/05/2026", "url": "https://eba.eu/1", "includi_in_pptx": "SI"},
    {"categoria": "APPROFONDIMENTI", "fonte": "EIOPA",
     "titolo": "Studio EIOPA", "descrizione": "Studio pubblicato.",
     "data_originale": "13/05/2026", "url": "https://eiopa.eu/1", "includi_in_pptx": "NO"},
]

TEMPLATE_PATH = "assets/Template_settimanale.xlsx"

# Mese e anno "odierni" per i test dei filtri K/L
TODAY_MESE = date.today().strftime("%m")
TODAY_ANNO = date.today().strftime("%Y")


@pytest.fixture
def output_path(tmp_path):
    if not os.path.exists(TEMPLATE_PATH):
        pytest.skip("Template XLSX non trovato in assets/")
    dest = str(tmp_path / "test_output.xlsx")
    shutil.copy(TEMPLATE_PATH, dest)
    return dest


def test_get_output_path_format():
    path = get_output_path()
    assert path == "output/DB_EXCEL/alert_normativo_DB.xlsx"


def test_ensure_excel_exists_creates_file(tmp_path):
    """File mancante → viene creato, restituisce True."""
    if not os.path.exists(TEMPLATE_PATH):
        pytest.skip("Template XLSX non trovato")
    dest = str(tmp_path / "out.xlsx")
    result = ensure_excel_exists(TEMPLATE_PATH, dest)
    assert result is True
    assert os.path.exists(dest)


def test_ensure_excel_exists_does_not_overwrite(tmp_path):
    """File già esistente → non viene toccato, restituisce False."""
    if not os.path.exists(TEMPLATE_PATH):
        pytest.skip("Template XLSX non trovato")
    dest = str(tmp_path / "out.xlsx")
    shutil.copy(TEMPLATE_PATH, dest)
    mtime_before = os.path.getmtime(dest)
    result = ensure_excel_exists(TEMPLATE_PATH, dest)
    assert result is False
    assert os.path.getmtime(dest) == mtime_before


def test_count_existing_rows_empty(tmp_path):
    """Template appena copiato (nessuna riga dati) → 0."""
    if not os.path.exists(TEMPLATE_PATH):
        pytest.skip("Template XLSX non trovato")
    dest = str(tmp_path / "out.xlsx")
    shutil.copy(TEMPLATE_PATH, dest)
    assert count_existing_rows(dest) == 0


def test_count_existing_rows_after_append(output_path):
    """Dopo append_news con 2 notizie → count_existing_rows restituisce 2."""
    append_news(output_path, SAMPLE_NEWS)
    assert count_existing_rows(output_path) == 2


def test_append_news_writes_rows(output_path):
    added = append_news(output_path, SAMPLE_NEWS)
    assert added == 2
    wb = load_workbook(output_path)
    ws = wb["Monitoraggio finance"]
    urls = [ws.cell(row=r, column=9).value for r in range(3, ws.max_row + 1)
            if ws.cell(row=r, column=1).value is not None]
    assert "https://eba.eu/1" in urls
    assert "https://eiopa.eu/1" in urls


def test_append_news_populates_kl_columns(output_path):
    """append_news deve scrivere MM in col K e AAAA in col L per ogni riga."""
    append_news(output_path, SAMPLE_NEWS)
    wb = load_workbook(output_path)
    ws = wb["Monitoraggio finance"]
    for r in range(3, ws.max_row + 1):
        if ws.cell(r, 1).value is None:
            continue
        assert ws.cell(r, 11).value == TODAY_MESE, f"riga {r}: col K attesa {TODAY_MESE!r}"
        assert ws.cell(r, 12).value == TODAY_ANNO, f"riga {r}: col L attesa {TODAY_ANNO!r}"


def test_append_news_deduplicates_by_url(output_path):
    append_news(output_path, SAMPLE_NEWS)
    added_second = append_news(output_path, SAMPLE_NEWS)
    assert added_second == 0
    wb = load_workbook(output_path)
    ws = wb["Monitoraggio finance"]
    urls = [ws.cell(row=r, column=9).value for r in range(3, ws.max_row + 1)
            if ws.cell(row=r, column=1).value is not None]
    assert urls.count("https://eba.eu/1") == 1


def test_append_news_returns_count(output_path):
    added = append_news(output_path, SAMPLE_NEWS)
    assert added == 2


# ── Tests for backup rolling ──────────────────────────────────────────────────

def _backup_path_for(path: str) -> str:
    return os.path.splitext(path)[0] + ".backup.xlsx"


def test_append_news_creates_backup(output_path):
    """append_news crea il backup con lo stato PRE-scrittura (template vuoto → 0 righe)."""
    append_news(output_path, SAMPLE_NEWS)
    backup = _backup_path_for(output_path)
    assert os.path.exists(backup)
    assert count_existing_rows(backup) == 0


def test_backup_preserves_previous_state(output_path):
    """Il backup contiene sempre lo stato della run precedente, non quello corrente."""
    append_news(output_path, SAMPLE_NEWS)
    nuova = [{"categoria": "CROSS FINANCE", "fonte": "ESMA",
              "titolo": "ESMA aggiorna Q&A MiFID II", "descrizione": "Aggiornamento Q&A.",
              "data_originale": "15/05/2026", "url": "https://esma.eu/1", "includi_in_pptx": "SI"}]
    append_news(output_path, nuova)
    backup = _backup_path_for(output_path)
    assert count_existing_rows(backup) == 2      # stato pre-seconda-scrittura
    assert count_existing_rows(output_path) == 3  # file principale aggiornato


# ── Helpers for read_approved_news tests ──────────────────────────────────────

def _write_filter_columns(path: str, edizione: str, mese: str, anno: str, start_row: int = 3):
    """Write edizione (col J), mese (col K) and anno (col L) for all data rows."""
    wb = openpyxl.load_workbook(path)
    ws = wb["Monitoraggio finance"]
    for row_num in range(start_row, ws.max_row + 1):
        if ws.cell(row_num, 1).value is not None:
            ws.cell(row_num, 10).value = edizione   # column J
            ws.cell(row_num, 11).value = mese        # column K
            ws.cell(row_num, 12).value = anno        # column L
    wb.save(path)


# ── Tests for read_approved_news ──────────────────────────────────────────────

def test_read_approved_news_returns_si_only(output_path):
    append_news(output_path, SAMPLE_NEWS)
    _write_filter_columns(output_path, edizione="1", mese=TODAY_MESE, anno=TODAY_ANNO)
    result = read_approved_news(output_path, edizione_numero="1", mese=TODAY_MESE, anno=TODAY_ANNO)
    # SAMPLE_NEWS[0]: includi_in_pptx=SI → included
    # SAMPLE_NEWS[1]: includi_in_pptx=NO → excluded
    urls = [item["url"] for items in result.values() for item in items]
    assert "https://eba.eu/1" in urls
    assert "https://eiopa.eu/1" not in urls


def test_read_approved_news_filters_by_edition(output_path):
    append_news(output_path, SAMPLE_NEWS)
    _write_filter_columns(output_path, edizione="2", mese=TODAY_MESE, anno=TODAY_ANNO)
    result = read_approved_news(output_path, edizione_numero="1", mese=TODAY_MESE, anno=TODAY_ANNO)
    urls = [item["url"] for items in result.values() for item in items]
    assert urls == []


def test_read_approved_news_filters_by_mese(output_path):
    """Notizie con mese diverso da quello richiesto non devono essere incluse."""
    append_news(output_path, SAMPLE_NEWS)
    _write_filter_columns(output_path, edizione="1", mese="01", anno=TODAY_ANNO)
    result = read_approved_news(output_path, edizione_numero="1", mese="99", anno=TODAY_ANNO)
    urls = [item["url"] for items in result.values() for item in items]
    assert urls == []


def test_read_approved_news_filters_by_anno(output_path):
    """Notizie con anno diverso da quello richiesto non devono essere incluse."""
    append_news(output_path, SAMPLE_NEWS)
    _write_filter_columns(output_path, edizione="1", mese=TODAY_MESE, anno="2000")
    result = read_approved_news(output_path, edizione_numero="1", mese=TODAY_MESE, anno=TODAY_ANNO)
    urls = [item["url"] for items in result.values() for item in items]
    assert urls == []


def test_read_approved_news_groups_by_category(output_path):
    append_news(output_path, SAMPLE_NEWS)
    _write_filter_columns(output_path, edizione="1", mese=TODAY_MESE, anno=TODAY_ANNO)
    result = read_approved_news(output_path, edizione_numero="1", mese=TODAY_MESE, anno=TODAY_ANNO)
    assert "BANKING" in result
    assert isinstance(result["BANKING"], list)
    assert result["BANKING"][0]["url"] == "https://eba.eu/1"


def test_read_approved_news_no_mese_anno_filter(output_path):
    """Se mese e anno sono omessi, il filtro K/L non si applica."""
    append_news(output_path, SAMPLE_NEWS)
    _write_filter_columns(output_path, edizione="1", mese="01", anno="2000")
    result = read_approved_news(output_path, edizione_numero="1")   # no mese/anno
    urls = [item["url"] for items in result.values() for item in items]
    assert "https://eba.eu/1" in urls


# ── Tests for count_filter_stages ─────────────────────────────────────────────

def test_count_filter_stages_full_match(output_path):
    """SAMPLE_NEWS: 2 righe, 1 approvata (SI); J/L corrette → tutti gli stadi pieni."""
    append_news(output_path, SAMPLE_NEWS)
    _write_filter_columns(output_path, edizione="1", mese=TODAY_MESE, anno=TODAY_ANNO)
    stages = count_filter_stages(output_path, edizione_numero="1", anno=TODAY_ANNO)
    assert stages == {"righe": 2, "approvate": 1, "edizione": 1, "anno": 1}


def test_count_filter_stages_wrong_edition(output_path):
    """Edizione non corrispondente → lo stadio 'edizione' va a zero."""
    append_news(output_path, SAMPLE_NEWS)
    _write_filter_columns(output_path, edizione="2", mese=TODAY_MESE, anno=TODAY_ANNO)
    stages = count_filter_stages(output_path, edizione_numero="1", anno=TODAY_ANNO)
    assert stages["approvate"] == 1
    assert stages["edizione"] == 0
    assert stages["anno"] == 0


def test_count_filter_stages_wrong_anno(output_path):
    """Anno non corrispondente → 'edizione' pieno ma 'anno' a zero."""
    append_news(output_path, SAMPLE_NEWS)
    _write_filter_columns(output_path, edizione="1", mese=TODAY_MESE, anno="2000")
    stages = count_filter_stages(output_path, edizione_numero="1", anno=TODAY_ANNO)
    assert stages["edizione"] == 1
    assert stages["anno"] == 0
