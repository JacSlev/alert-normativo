import shutil
import os
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Template structure:
# Row 1: title ("MONITORAGGIO FINANCE")
# Row 2: column headers (ID, Macro-Ambito, ..., Link)
# Row 3+: data rows

HEADER_ROW = 2
DATA_START_ROW = 3

CATEGORY_ORDER = ["BANKING", "INSURANCE", "CROSS FINANCE", "APPROFONDIMENTI"]


def get_output_path() -> str:
    return "output/DB_EXCEL/alert_normativo_DB.xlsx"


def ensure_excel_exists(template_path: str, output_path: str) -> bool:
    """Copy template to output_path if it does not already exist.

    Returns True if the file was created, False if it already existed.
    """
    if os.path.exists(output_path):
        return False
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    shutil.copy(template_path, output_path)
    return True


def count_existing_rows(output_path: str) -> int:
    """Return the number of data rows (column A not None) in 'Monitoraggio finance'."""
    wb = load_workbook(output_path, read_only=True)
    ws = wb["Monitoraggio finance"]
    count = 0
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if row[0] is not None:
            count += 1
    wb.close()
    return count


def _get_existing_urls(ws) -> set:
    urls = set()
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if row[8]:  # column I (0-indexed: 8)
            urls.add(row[8])
    return urls


def _next_id(ws) -> int:
    max_id = 0
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        val = row[0]  # column A
        if val is not None:
            try:
                max_id = max(max_id, int(val))
            except (ValueError, TypeError):
                pass
    return max_id + 1


def _first_empty_row(ws) -> int:
    """Return the row number of the first row >= DATA_START_ROW where column A is None."""
    for row_num in range(DATA_START_ROW, ws.max_row + 2):
        if ws.cell(row_num, 1).value is None:
            return row_num
    return ws.max_row + 1


def _format_sheet(ws) -> None:
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_font = Font(bold=True)
    for cell in ws[HEADER_ROW]:
        cell.fill = header_fill
        cell.font = header_font

    ws.auto_filter.ref = ws.dimensions

    for col_idx, col_cells in enumerate(ws.columns, 1):
        col_letter = get_column_letter(col_idx)
        if col_idx == 7:  # column G — description, fixed width + wrap
            ws.column_dimensions[col_letter].width = 60
            for cell in col_cells[DATA_START_ROW - 1:]:
                cell.alignment = Alignment(wrap_text=True)
        else:
            max_len = max(
                (len(str(c.value)) for c in col_cells if c.value is not None),
                default=10,
            )
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)


def append_news(output_path: str, news_items: list[dict]) -> int:
    wb = load_workbook(output_path)
    ws = wb["Monitoraggio finance"]

    existing_urls = _get_existing_urls(ws)
    next_id = _next_id(ws)
    today = date.today().strftime("%d/%m/%Y")
    row_num = _first_empty_row(ws)
    added = 0

    for item in news_items:
        url = item.get("url", "")
        if url in existing_urls:
            print(f"[SKIP] Duplicato: {url}")
            continue
        values = [
            next_id,
            item.get("categoria", ""),
            item.get("fonte", ""),
            item.get("data_originale", ""),
            today,
            item.get("titolo", ""),
            item.get("descrizione", ""),
            item.get("includi_in_pptx", "NO"),
            url,
        ]
        for col_idx, value in enumerate(values, 1):
            ws.cell(row_num, col_idx).value = value
        # Colonna K (11) = mese odierno MM, Colonna L (12) = anno odierno AAAA
        ws.cell(row_num, 11).value = date.today().strftime("%m")
        ws.cell(row_num, 12).value = date.today().strftime("%Y")
        existing_urls.add(url)
        next_id += 1
        row_num += 1
        added += 1

    _format_sheet(ws)
    wb.save(output_path)
    return added


def read_approved_news(excel_path: str, edizione_numero: str, mese: str = "", anno: str = "") -> dict:
    """Read the reviewed Excel file and return approved news grouped by category.

    Filters:
    - Column H (index 7)  = "SI"            — approved by the reviewer
    - Column J (index 9)  = edizione_numero — belongs to this edition
    - Column K (index 10) = mese (MM)       — if non-empty, filters by month
    - Column L (index 11) = anno (AAAA)     — if non-empty, filters by year

    Returns a dict keyed by category name (CATEGORY_ORDER), each value a list of:
        {"descrizione": str, "data": str, "url": str, "fonte": str}
    """
    wb = load_workbook(excel_path)
    ws = wb["Monitoraggio finance"]
    grouped: dict = {cat: [] for cat in CATEGORY_ORDER}

    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if row[0] is None:
            continue
        includi = str(row[7]).strip().upper() if row[7] else "NO"
        if includi != "SI":
            continue
        edizione = str(row[9]).strip() if len(row) > 9 and row[9] is not None else ""
        if edizione != str(edizione_numero).strip():
            continue
        if mese:
            riga_mese = str(row[10]).strip() if len(row) > 10 and row[10] is not None else ""
            if riga_mese != mese.strip():
                continue
        if anno:
            riga_anno = str(row[11]).strip() if len(row) > 11 and row[11] is not None else ""
            if riga_anno != anno.strip():
                continue
        categoria = str(row[1]).strip() if row[1] else ""
        if categoria not in grouped:
            grouped[categoria] = []
        grouped[categoria].append({
            "descrizione": row[6] or "",
            "data":        row[3] or "",
            "url":         row[8] or "",
            "fonte":       row[2] or "",
        })
    return grouped
